from __future__ import annotations

import io
from urllib.parse import quote

import requests
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from . import aura_agent_core as core
from . import aura_agent_tools as tools
from .aura_connector_hardening import install_aura_connector_hardening
from .aura_connectors import google

_INSTALLED = False
_MAX_BYTES = 8 * 1024 * 1024
_MAX_TEXT = 80_000
_GOOGLE_DOC = "application/vnd.google-apps.document"
_GOOGLE_SHEET = "application/vnd.google-apps.spreadsheet"
_GOOGLE_SLIDES = "application/vnd.google-apps.presentation"
_GOOGLE_SCRIPT = "application/vnd.google-apps.script"
_DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


READ_SPEC = tools.ToolSpec(
    "google_drive_read_text",
    "Read bounded text from one file in the member's connected Google Drive. Supports Google Docs/Sheets/Slides, text, PDF, DOCX and XLSX. Read-only; does not add the file to a Studio project.",
    {"file_id": "Google Drive file id returned by Drive search."},
)


def _download(user_id: str, url: str, *, params: dict | None = None, max_bytes: int = _MAX_BYTES) -> tuple[bytes, str]:
    token = google._token(user_id, "drive")
    response = requests.get(
        url,
        params=params or {},
        headers={"Authorization": f"Bearer {token}"},
        timeout=google.timeout,
        stream=True,
    )
    response.raise_for_status()
    data = bytearray()
    for chunk in response.iter_content(chunk_size=256 * 1024):
        if not chunk:
            continue
        data.extend(chunk)
        if len(data) > max_bytes:
            raise ValueError(f"Drive file exceeds Aura's {max_bytes // (1024 * 1024)} MB connector read limit")
    return bytes(data), str(response.headers.get("content-type") or "").split(";", 1)[0].lower()


def _metadata(user_id: str, file_id: str) -> dict:
    clean = (file_id or "").strip()
    if not clean or len(clean) > 300:
        raise ValueError("A valid Google Drive file id is required")
    return google._get(
        user_id,
        "drive",
        f"https://www.googleapis.com/drive/v3/files/{quote(clean, safe='')}",
        params={"fields": "id,name,mimeType,size,modifiedTime,webViewLink,capabilities(canDownload)"},
    )


def _decode_text(data: bytes) -> str:
    return data.decode("utf-8-sig", errors="replace")


def _docx_text(data: bytes) -> str:
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _xlsx_text(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        rows = []
        chars = 0
        for ws in wb.worksheets[:12]:
            heading = f"[Sheet: {ws.title}]"
            rows.append(heading)
            chars += len(heading)
            for row in ws.iter_rows(max_row=300, values_only=True):
                values = ["" if value is None else str(value) for value in row[:40]]
                if any(values):
                    line = "\t".join(values)
                    rows.append(line)
                    chars += len(line)
                if chars > _MAX_TEXT * 2:
                    break
            if chars > _MAX_TEXT * 2:
                break
        return "\n".join(rows)
    finally:
        wb.close()


def _pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages[:100])


def read_drive_text(user_id: str, file_id: str) -> dict:
    meta = _metadata(user_id, file_id)
    mime = str(meta.get("mimeType") or "").lower()
    can_download = bool((meta.get("capabilities") or {}).get("canDownload", True))
    if not can_download:
        raise PermissionError("This Drive file does not allow downloading/exporting")

    file_url = f"https://www.googleapis.com/drive/v3/files/{quote(str(meta.get('id') or file_id), safe='')}"
    exported_as = None
    if mime == _GOOGLE_DOC:
        data, content_type = _download(user_id, file_url + "/export", params={"mimeType": "text/plain"})
        text, exported_as = _decode_text(data), "text/plain"
    elif mime == _GOOGLE_SHEET:
        data, content_type = _download(user_id, file_url + "/export", params={"mimeType": "text/csv"})
        text, exported_as = _decode_text(data), "text/csv:first-sheet"
    elif mime == _GOOGLE_SLIDES:
        data, content_type = _download(user_id, file_url + "/export", params={"mimeType": "text/plain"})
        text, exported_as = _decode_text(data), "text/plain"
    elif mime == _GOOGLE_SCRIPT:
        data, content_type = _download(user_id, file_url + "/export", params={"mimeType": "application/vnd.google-apps.script+json"})
        text, exported_as = _decode_text(data), "application/vnd.google-apps.script+json"
    else:
        data, content_type = _download(user_id, file_url, params={"alt": "media"})
        if mime.startswith("text/") or mime in {"application/json", "application/xml", "application/yaml", "application/x-yaml"}:
            text = _decode_text(data)
        elif mime == "application/pdf" or content_type == "application/pdf":
            text = _pdf_text(data)
        elif mime == _DOCX:
            text = _docx_text(data)
        elif mime in {_XLSX, "application/vnd.ms-excel"}:
            text = _xlsx_text(data)
        else:
            raise ValueError(f"Aura Drive reading does not yet extract this file type: {mime or content_type or 'unknown'}")

    clean = text.strip()
    return {
        "file": {
            "id": meta.get("id"),
            "name": meta.get("name"),
            "mime_type": mime,
            "modified_time": meta.get("modifiedTime"),
            "web_view_link": meta.get("webViewLink"),
        },
        "text": clean[:_MAX_TEXT],
        "characters": len(clean),
        "truncated": len(clean) > _MAX_TEXT,
        "exported_as": exported_as,
        "read_only": True,
        "added_to_project": False,
    }


def _wants_drive_read(text: str) -> bool:
    lower = (text or "").lower()
    return any(term in lower for term in ("read it", "read the", "open it", "open the", "summarize it", "summarise it", "summarize the", "summarise the"))


def _drive_search_query(text: str) -> str | None:
    lower = text.lower()
    for phrase in ("search my drive for", "search google drive for", "find in my drive", "find in google drive"):
        index = lower.find(phrase)
        if index < 0:
            continue
        value = text[index + len(phrase):].strip()
        # Remove common tail instructions so the Drive fullText query stays focused.
        value = value.split(" and read", 1)[0].split(" and open", 1)[0].split(" and summarize", 1)[0].split(" and summarise", 1)[0]
        return value.strip(" :,-.?") or None
    return None


def install_aura_connector_extensions() -> None:
    global _INSTALLED
    if _INSTALLED:
        return
    install_aura_connector_hardening()
    if READ_SPEC.name not in {item.name for item in tools.TOOL_SPECS}:
        tools.TOOL_SPECS.append(READ_SPEC)
        tools._SPEC_BY_NAME[READ_SPEC.name] = READ_SPEC

    original_execute = tools.AuraToolRegistry.execute
    original_direct = core._direct_tool_plan
    original_needs = core._needs_model_tool_router

    def execute(self, call: tools.ToolCall, *, latest_user_message: str):
        if call.name != "google_drive_read_text":
            return original_execute(self, call, latest_user_message=latest_user_message)
        if not self.tools_enabled:
            raise PermissionError("Aura tools are disabled for this conversation")
        return read_drive_text(self.member.user_id, str((call.arguments or {}).get("file_id") or ""))

    def direct_tool_plan(text: str, pinned_project: str | None, web_enabled: bool):
        query = _drive_search_query(text)
        if query and _wants_drive_read(text):
            return tools.ToolPlan(calls=[
                tools.ToolCall(name="google_drive_search", arguments={"query": query, "limit": 10}),
                tools.ToolCall(name="google_drive_read_text", arguments={"file_id": "$step0.files.0.id"}),
            ])
        return original_direct(text, pinned_project, web_enabled)

    def needs_model_tool_router(text: str, pinned_project: str | None, tools_enabled: bool, web_enabled: bool) -> bool:
        if tools_enabled and "drive" in (text or "").lower() and _wants_drive_read(text):
            return True
        return original_needs(text, pinned_project, tools_enabled, web_enabled)

    tools.AuraToolRegistry.execute = execute
    core._direct_tool_plan = direct_tool_plan
    core._needs_model_tool_router = needs_model_tool_router
    _INSTALLED = True


__all__ = ["install_aura_connector_extensions", "read_drive_text", "_drive_search_query"]
