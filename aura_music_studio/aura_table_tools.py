from __future__ import annotations

import csv
import math
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from uuid import uuid4
from xml.sax.saxutils import escape as xml_escape

from openpyxl import load_workbook

from . import aura_agent_core as core
from . import aura_agent_tools as tools
from .creative_project import CreativeProjectStore

_INSTALLED = False
_TABLE_EXTS = {".csv", ".tsv", ".xlsx", ".xlsm"}
_MAX_ROWS = 50_000
_MAX_COLS = 200


def _spec(name: str, description: str, arguments: dict[str, str], *, write: bool = False):
    return tools.ToolSpec(name=name, description=description, arguments=arguments, write=write, web=False)


TABLE_SPECS = [
    _spec(
        "inspect_project_table",
        "Inspect a CSV/TSV/XLSX already stored in the pinned project: sheets, dimensions, columns, types and a bounded preview.",
        {"project_name": "Project name; omit if pinned.", "table": "Optional filename/label/source selector.", "sheet": "Optional XLSX sheet name."},
    ),
    _spec(
        "analyze_project_table",
        "Safely profile a project CSV/TSV/XLSX: missing values, numeric statistics, top categorical values and bounded correlations. No arbitrary code execution.",
        {"project_name": "Project name; omit if pinned.", "table": "Optional filename/label selector.", "sheet": "Optional XLSX sheet name."},
    ),
    _spec(
        "group_project_table",
        "Group a project table by one column and return counts plus optional sum/mean for a numeric value column.",
        {"project_name": "Project name.", "table": "Optional table selector.", "sheet": "Optional sheet.", "group_column": "Column name.", "value_column": "Optional numeric column."},
    ),
    _spec(
        "chart_project_table",
        "Create a bounded SVG bar or line chart from a project table and save it as a project output. Use only when the member explicitly asks for a chart/plot/graph.",
        {"project_name": "Project name.", "table": "Optional table selector.", "sheet": "Optional sheet.", "x_column": "Category/x column.", "y_column": "Numeric y column.", "chart_type": "bar or line.", "limit": "Optional 2-40 rows, default 20."},
        write=True,
    ),
]


def _safe_project_source(project: Path, source_ref: str) -> Path:
    root = project.resolve()
    target = (root / source_ref).resolve()
    if target != root and root not in target.parents:
        raise PermissionError("Table source resolves outside the member project")
    if not target.is_file() or target.suffix.lower() not in _TABLE_EXTS:
        raise FileNotFoundError("Project table source is unavailable")
    return target


def _candidates(project: Path) -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()
    creative = CreativeProjectStore(project)
    if creative.exists():
        try:
            manifest = creative.load()
            for ref in manifest.references:
                if Path(ref.source_ref).suffix.lower() not in _TABLE_EXTS:
                    continue
                try:
                    path = _safe_project_source(project, ref.source_ref)
                except Exception:
                    continue
                key = path.as_posix()
                if key in seen:
                    continue
                seen.add(key)
                rows.append({"label": ref.label, "source_ref": ref.source_ref, "path": path})
        except Exception:
            pass
    for folder in (project / "input" / "references", project / "input"):
        if not folder.is_dir():
            continue
        for path in folder.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in _TABLE_EXTS:
                continue
            resolved = path.resolve()
            key = resolved.as_posix()
            if key in seen:
                continue
            seen.add(key)
            rows.append({"label": path.name, "source_ref": resolved.relative_to(project.resolve()).as_posix(), "path": resolved})
    return rows


def _select_table(project: Path, selector: str | None) -> dict:
    rows = _candidates(project)
    if not rows:
        raise FileNotFoundError("No CSV/TSV/XLSX table is registered in this project")
    clean = (selector or "").strip().lower()
    if not clean:
        if len(rows) == 1:
            return rows[0]
        raise ValueError("Multiple project tables are available: " + ", ".join(row["label"] for row in rows[:12]))
    exact = [row for row in rows if clean in {str(row["label"]).lower(), Path(row["source_ref"]).name.lower(), str(row["source_ref"]).lower()}]
    if len(exact) == 1:
        return exact[0]
    partial = [row for row in rows if clean in str(row["label"]).lower() or clean in str(row["source_ref"]).lower()]
    if len(partial) == 1:
        return partial[0]
    if not partial:
        raise KeyError(f"No project table matches {selector!r}")
    raise ValueError("Table selector is ambiguous: " + ", ".join(row["label"] for row in partial[:12]))


def _clean_header(value: Any, index: int) -> str:
    text = " ".join(str(value if value is not None else "").split()).strip()
    return text[:160] or f"column_{index + 1}"


def _read_csv(path: Path) -> tuple[list[str], list[list[Any]], str]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        if path.suffix.lower() == ".csv":
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except Exception:
                pass
        reader = csv.reader(handle, delimiter=delimiter)
        first = next(reader, [])[:_MAX_COLS]
        headers = [_clean_header(value, index) for index, value in enumerate(first)]
        rows: list[list[Any]] = []
        for index, row in enumerate(reader):
            if index >= _MAX_ROWS:
                break
            rows.append(list(row[: len(headers)]))
    return headers, rows, f"delimiter:{delimiter}"


def _read_excel(path: Path, sheet: str | None) -> tuple[list[str], list[list[Any]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = sheet if sheet and sheet in workbook.sheetnames else workbook.sheetnames[0]
        if sheet and sheet not in workbook.sheetnames:
            raise KeyError("Unknown XLSX sheet. Available: " + ", ".join(workbook.sheetnames[:30]))
        ws = workbook[selected]
        iterator = ws.iter_rows(values_only=True)
        first = list(next(iterator, ()))[:_MAX_COLS]
        headers = [_clean_header(value, index) for index, value in enumerate(first)]
        rows: list[list[Any]] = []
        for index, row in enumerate(iterator):
            if index >= _MAX_ROWS:
                break
            rows.append(list(row[: len(headers)]))
        return headers, rows, selected
    finally:
        workbook.close()


def _read_table(path: Path, sheet: str | None = None) -> tuple[list[str], list[list[Any]], str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_excel(path, sheet)
    return _read_csv(path)


def _to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        try:
            number = float(text[:-1]) / 100.0
            return number if math.isfinite(number) else None
        except ValueError:
            return None
    try:
        number = float(text)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def _column_values(headers: list[str], rows: list[list[Any]]) -> dict[str, list[Any]]:
    values = {header: [] for header in headers}
    for row in rows:
        for index, header in enumerate(headers):
            values[header].append(row[index] if index < len(row) else None)
    return values


def _numeric_column(values: list[Any]) -> list[float] | None:
    populated = [value for value in values if value not in {None, ""}]
    if not populated:
        return None
    converted = [_to_number(value) for value in populated]
    numeric = [value for value in converted if value is not None]
    if len(numeric) / len(populated) < 0.8:
        return None
    return numeric


def _profile(headers: list[str], rows: list[list[Any]]) -> dict:
    columns = _column_values(headers, rows)
    result: dict[str, dict] = {}
    numeric_vectors: dict[str, list[float]] = {}
    for header, values in columns.items():
        missing = sum(value is None or str(value).strip() == "" for value in values)
        numeric = _numeric_column(values)
        if numeric:
            numeric_vectors[header] = numeric
            item = {
                "type": "numeric",
                "count_numeric": len(numeric),
                "missing": missing,
                "mean": statistics.fmean(numeric),
                "median": statistics.median(numeric),
                "min": min(numeric),
                "max": max(numeric),
            }
            if len(numeric) >= 2:
                item["sample_stdev"] = statistics.stdev(numeric)
            result[header] = item
        else:
            strings = [str(value).strip() for value in values if value is not None and str(value).strip()]
            counts = Counter(strings)
            result[header] = {
                "type": "categorical_or_text",
                "count_populated": len(strings),
                "missing": missing,
                "unique": len(counts),
                "top_values": counts.most_common(8),
            }
    correlations = []
    numeric_names = list(numeric_vectors)[:20]
    # Pair rows rather than independently compressed columns so missing values remain aligned.
    for i, left in enumerate(numeric_names):
        li = headers.index(left)
        for right in numeric_names[i + 1 :]:
            ri = headers.index(right)
            pairs = []
            for row in rows:
                if li >= len(row) or ri >= len(row):
                    continue
                a, b = _to_number(row[li]), _to_number(row[ri])
                if a is not None and b is not None:
                    pairs.append((a, b))
            if len(pairs) < 3:
                continue
            xs = [a for a, _ in pairs]
            ys = [b for _, b in pairs]
            mean_x, mean_y = statistics.fmean(xs), statistics.fmean(ys)
            numerator = sum((x - mean_x) * (y - mean_y) for x, y in pairs)
            denom_x = math.sqrt(sum((x - mean_x) ** 2 for x in xs))
            denom_y = math.sqrt(sum((y - mean_y) ** 2 for y in ys))
            if denom_x and denom_y:
                corr = numerator / (denom_x * denom_y)
                if abs(corr) >= 0.25:
                    correlations.append({"left": left, "right": right, "pearson": round(corr, 5), "paired_rows": len(pairs)})
    correlations.sort(key=lambda item: abs(item["pearson"]), reverse=True)
    return {"columns": result, "correlations": correlations[:30]}


def _resolve_column(headers: list[str], selector: str) -> str:
    clean = (selector or "").strip().lower()
    exact = [header for header in headers if header.lower() == clean]
    if len(exact) == 1:
        return exact[0]
    partial = [header for header in headers if clean and clean in header.lower()]
    if len(partial) == 1:
        return partial[0]
    if not partial:
        raise KeyError(f"Unknown table column: {selector}")
    raise ValueError("Column selector is ambiguous: " + ", ".join(partial[:12]))


def _svg_chart(headers: list[str], rows: list[list[Any]], x_name: str, y_name: str, chart_type: str, limit: int) -> str:
    xi, yi = headers.index(x_name), headers.index(y_name)
    points = []
    for row in rows:
        if xi >= len(row) or yi >= len(row):
            continue
        y = _to_number(row[yi])
        if y is None:
            continue
        points.append((str(row[xi])[:50], y))
        if len(points) >= limit:
            break
    if not points:
        raise ValueError("No numeric chart rows are available")
    width, height, left, top, bottom = 1000, 600, 90, 60, 100
    inner_w, inner_h = width - left - 40, height - top - bottom
    values = [value for _, value in points]
    lo, hi = min(0.0, min(values)), max(0.0, max(values))
    span = hi - lo or 1.0
    def y_px(value):
        return top + (hi - value) / span * inner_h
    zero_y = y_px(0.0)
    chunks = [f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">', '<rect width="100%" height="100%" fill="#080b16"/>', f'<text x="{left}" y="30" fill="#ffffff" font-family="sans-serif" font-size="22">{xml_escape(y_name)} by {xml_escape(x_name)}</text>', f'<line x1="{left}" y1="{zero_y:.2f}" x2="{left+inner_w}" y2="{zero_y:.2f}" stroke="#8290aa" stroke-width="1"/>']
    step = inner_w / max(1, len(points))
    coords = []
    for index, (label, value) in enumerate(points):
        cx = left + step * (index + 0.5)
        yp = y_px(value)
        coords.append((cx, yp))
        if chart_type == "line":
            chunks.append(f'<circle cx="{cx:.2f}" cy="{yp:.2f}" r="4" fill="#edcc73"/>')
        else:
            bar_w = max(3.0, step * 0.64)
            y0, y1 = sorted((zero_y, yp))
            chunks.append(f'<rect x="{cx-bar_w/2:.2f}" y="{y0:.2f}" width="{bar_w:.2f}" height="{max(1,y1-y0):.2f}" rx="3" fill="#9b70ff"/>')
        if len(points) <= 24:
            chunks.append(f'<text x="{cx:.2f}" y="{height-55}" fill="#c8d0e3" font-family="sans-serif" font-size="12" text-anchor="middle" transform="rotate(-35 {cx:.2f} {height-55})">{xml_escape(label)}</text>')
    if chart_type == "line" and len(coords) > 1:
        path = " ".join(("M" if i == 0 else "L") + f" {x:.2f} {y:.2f}" for i, (x, y) in enumerate(coords))
        chunks.insert(4, f'<path d="{path}" fill="none" stroke="#58dfff" stroke-width="3"/>')
    chunks.append('</svg>')
    return "".join(chunks)


def _project_and_table(registry, args: dict):
    name = tools._project_name(args, registry.pinned_project)
    project = tools._safe_project(name)
    selected = _select_table(project, str(args.get("table") or "") or None)
    headers, rows, sheet = _read_table(selected["path"], str(args.get("sheet") or "") or None)
    if not headers:
        raise ValueError("Table has no header row")
    return name, project, selected, headers, rows, sheet


def _explicit_chart(text: str) -> bool:
    lower = (text or "").lower()
    return any(word in lower for word in ("chart", "plot", "graph", "visualize", "visualise"))


def _looks_table_request(text: str) -> bool:
    lower = (text or "").lower()
    return any(word in lower for word in ("spreadsheet", "xlsx", "csv", "table", "dataset", "data file")) and any(word in lower for word in ("analyze", "analyse", "inspect", "summarize", "summarise", "statistics", "profile", "group", "chart", "plot", "graph"))


def install_aura_table_tools() -> None:
    global _INSTALLED
    if _INSTALLED:
        return
    for spec in TABLE_SPECS:
        if spec.name not in {item.name for item in tools.TOOL_SPECS}:
            tools.TOOL_SPECS.append(spec)
            tools._SPEC_BY_NAME[spec.name] = spec
    original_execute = tools.AuraToolRegistry.execute
    original_direct = core._direct_tool_plan
    original_needs = core._needs_model_tool_router
    names = {item.name for item in TABLE_SPECS}

    def execute(self, call: tools.ToolCall, *, latest_user_message: str):
        if call.name not in names:
            return original_execute(self, call, latest_user_message=latest_user_message)
        if not self.tools_enabled:
            raise PermissionError("Aura tools are disabled for this conversation")
        if call.name == "chart_project_table" and not _explicit_chart(latest_user_message):
            raise PermissionError("Creating a project chart requires an explicit chart/plot/graph request")
        args = dict(call.arguments or {})
        name, project, selected, headers, rows, sheet = _project_and_table(self, args)
        base = {"project_name": name, "table": selected["label"], "source_ref": selected["source_ref"], "sheet": sheet, "rows_loaded": len(rows), "columns": headers, "bounded": True, "code_execution": False}
        if call.name == "inspect_project_table":
            preview = [{headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))} for row in rows[:8]]
            profile = _profile(headers, rows[: min(len(rows), 5000)])
            base.update({"preview": preview, "column_types": {key: value["type"] for key, value in profile["columns"].items()}, "row_limit": _MAX_ROWS})
            return base
        if call.name == "analyze_project_table":
            base.update(_profile(headers, rows))
            return base
        if call.name == "group_project_table":
            group = _resolve_column(headers, str(args.get("group_column") or ""))
            value_selector = str(args.get("value_column") or "").strip()
            value = _resolve_column(headers, value_selector) if value_selector else None
            gi = headers.index(group)
            vi = headers.index(value) if value else None
            groups: dict[str, list[float]] = defaultdict(list)
            counts: Counter[str] = Counter()
            for row in rows:
                label = str(row[gi] if gi < len(row) else "").strip() or "(blank)"
                counts[label] += 1
                if vi is not None and vi < len(row):
                    number = _to_number(row[vi])
                    if number is not None:
                        groups[label].append(number)
            output = []
            for label, count in counts.most_common(100):
                item = {"group": label, "count": count}
                values = groups.get(label, [])
                if value and values:
                    item.update({"value_column": value, "sum": math.fsum(values), "mean": statistics.fmean(values), "numeric_count": len(values)})
                output.append(item)
            base.update({"group_column": group, "value_column": value, "groups": output})
            return base
        if call.name == "chart_project_table":
            x_name = _resolve_column(headers, str(args.get("x_column") or ""))
            y_name = _resolve_column(headers, str(args.get("y_column") or ""))
            chart_type = str(args.get("chart_type") or "bar").strip().lower()
            if chart_type not in {"bar", "line"}:
                raise ValueError("chart_type must be bar or line")
            limit = max(2, min(int(args.get("limit") or 20), 40))
            svg = _svg_chart(headers, rows, x_name, y_name, chart_type, limit)
            out_dir = project / "output" / "data"
            out_dir.mkdir(parents=True, exist_ok=True)
            stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(selected["label"]).stem).strip("._") or "table"
            output = out_dir / f"{stem}_{uuid4().hex[:10]}_{chart_type}.svg"
            output.write_text(svg, encoding="utf-8")
            base.update({"created": True, "chart_type": chart_type, "x_column": x_name, "y_column": y_name, "output": output.relative_to(project / "output").as_posix(), "storage_path_exposed": False})
            return base
        raise ValueError(f"Unsupported Aura table tool: {call.name}")

    def direct_tool_plan(text: str, pinned_project: str | None, web_enabled: bool):
        prior = original_direct(text, pinned_project, web_enabled)
        if prior is not None:
            return prior
        if not pinned_project or not _looks_table_request(text):
            return None
        lower = text.lower()
        if any(word in lower for word in ("chart", "plot", "graph", "visualize", "visualise")):
            # Column selection is best handled by the model router; no direct call without explicit columns.
            return None
        if "group" in lower:
            return None
        name = "inspect_project_table" if "inspect" in lower or "preview" in lower else "analyze_project_table"
        return tools.ToolPlan(calls=[tools.ToolCall(name=name, arguments={"project_name": pinned_project})])

    def needs_model_tool_router(text: str, pinned_project: str | None, tools_enabled: bool, web_enabled: bool) -> bool:
        if tools_enabled and pinned_project and _looks_table_request(text):
            return True
        return original_needs(text, pinned_project, tools_enabled, web_enabled)

    tools.AuraToolRegistry.execute = execute
    core._direct_tool_plan = direct_tool_plan
    core._needs_model_tool_router = needs_model_tool_router
    _INSTALLED = True


__all__ = ["install_aura_table_tools", "_profile", "_select_table"]
