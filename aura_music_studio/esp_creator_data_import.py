from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
from datetime import datetime, timezone
from html import escape
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator

from .esp_command_center import esp
from .esp_niche import require_esp_hub_member
from .esp_progress import EspProgressStore, save_progress_upload

router = APIRouter(tags=["ESP Creator Data Import"])
MAX_ROWS = 500
MAX_COLUMNS = 50
SUPPORTED_SUFFIXES = {".csv", ".json", ".xlsx"}

STANDARD_METRICS = (
    "views",
    "duration_minutes",
    "avg_watch_seconds",
    "completion_rate",
    "peak_viewers",
    "new_followers",
    "comments",
    "shares",
    "saves",
    "diamonds",
)

ALIASES = {
    "views": {"views", "view", "video_views", "total_views", "live_views", "viewers"},
    "duration_minutes": {"duration_minutes", "duration_mins", "live_duration_minutes", "minutes", "duration"},
    "avg_watch_seconds": {"avg_watch_seconds", "average_watch_seconds", "average_watch_time", "avg_watch_time", "watch_seconds"},
    "completion_rate": {"completion_rate", "completion_percent", "completion_percentage", "watched_full_video_percent"},
    "peak_viewers": {"peak_viewers", "peak_concurrent_viewers", "max_viewers", "peak_concurrent"},
    "new_followers": {"new_followers", "followers_gained", "follows", "new_follows"},
    "comments": {"comments", "comment_count", "total_comments"},
    "shares": {"shares", "share_count", "total_shares"},
    "saves": {"saves", "save_count", "favorites", "favourites"},
    "diamonds": {"diamonds", "diamond_count", "total_diamonds"},
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean(value: str | None, limit: int) -> str:
    return " ".join(str(value or "").split())[:limit]


def _header(value) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return text[:100]


def _number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value if value >= 0 else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        number = float(text)
    except ValueError:
        return None
    if number < 0:
        return None
    return int(number) if number.is_integer() else round(number, 4)


def _creator(request: Request):
    member, membership = require_esp_hub_member(request)
    role = "owner" if membership.get("status") == "owner" else (membership.get("roles") or "").lower()
    if role not in {"creator", "both", "owner"}:
        raise HTTPException(403, "ESP Creator access is required for performance data import")
    return member


class ImportConfirm(BaseModel):
    kind: str = Field(pattern=r"^(live|video)$")
    mapping: dict[str, str] = Field(default_factory=dict, max_length=20)
    period_column: str = Field(default="", max_length=100)
    default_period_label: str = Field(default="Imported analytics", max_length=160)
    notes: str = Field(default="", max_length=2000)

    @field_validator("mapping")
    @classmethod
    def validate_mapping(cls, value: dict[str, str]) -> dict[str, str]:
        result = {}
        for metric, column in value.items():
            if metric not in STANDARD_METRICS:
                raise ValueError(f"Unsupported progress metric: {metric}")
            clean = _header(column)
            if clean:
                result[metric] = clean
        return result


class CreatorDataImportStore:
    def __init__(self, db_path: str | None = None, progress_store: EspProgressStore | None = None):
        self.db_path = db_path or esp.db_path
        self.progress = progress_store or EspProgressStore(esp)
        self._init_schema()

    def _connect(self):
        con = sqlite3.connect(self.db_path)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA foreign_keys=ON")
        con.execute("PRAGMA journal_mode=WAL")
        return con

    def _init_schema(self) -> None:
        with self._connect() as con:
            con.executescript(
                """
                CREATE TABLE IF NOT EXISTS esp_creator_data_imports (
                    id TEXT PRIMARY KEY,
                    user_id TEXT NOT NULL,
                    upload_name TEXT NOT NULL,
                    upload_path TEXT NOT NULL,
                    upload_content_type TEXT NOT NULL DEFAULT '',
                    source_format TEXT NOT NULL,
                    columns_json TEXT NOT NULL DEFAULT '[]',
                    rows_json TEXT NOT NULL DEFAULT '[]',
                    detected_mapping_json TEXT NOT NULL DEFAULT '{}',
                    status TEXT NOT NULL DEFAULT 'staged',
                    imported_submission_ids_json TEXT NOT NULL DEFAULT '[]',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    imported_at TEXT,
                    rejected_at TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_creator_data_import_user
                    ON esp_creator_data_imports(user_id,status,created_at DESC);
                """
            )

    @staticmethod
    def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
        text = content.decode("utf-8-sig", errors="strict")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        raw_headers = reader.fieldnames or []
        headers = [_header(value) for value in raw_headers][:MAX_COLUMNS]
        if not headers or any(not value for value in headers):
            raise ValueError("CSV must contain a usable header row")
        if len(set(headers)) != len(headers):
            raise ValueError("CSV contains duplicate column names after normalisation")
        rows = []
        for raw in reader:
            row = {}
            for index, original in enumerate(raw_headers[:MAX_COLUMNS]):
                row[headers[index]] = str(raw.get(original) or "").strip()[:1000]
            if any(str(value).strip() for value in row.values()):
                rows.append(row)
            if len(rows) >= MAX_ROWS:
                break
        return headers, rows

    @staticmethod
    def _parse_json(content: bytes) -> tuple[list[str], list[dict]]:
        payload = json.loads(content.decode("utf-8-sig", errors="strict"))
        if isinstance(payload, dict):
            for key in ("rows", "data", "results", "records"):
                if isinstance(payload.get(key), list):
                    payload = payload[key]
                    break
            else:
                payload = [payload]
        if not isinstance(payload, list):
            raise ValueError("JSON must contain an object or an array of objects")
        raw_rows = [row for row in payload[:MAX_ROWS] if isinstance(row, dict)]
        if not raw_rows:
            raise ValueError("JSON does not contain any object rows")
        columns: list[str] = []
        seen = set()
        for raw in raw_rows:
            for key in raw:
                name = _header(key)
                if name and name not in seen:
                    columns.append(name)
                    seen.add(name)
                if len(columns) >= MAX_COLUMNS:
                    break
            if len(columns) >= MAX_COLUMNS:
                break
        rows = []
        for raw in raw_rows:
            normalized = {_header(key): value for key, value in raw.items() if _header(key)}
            row = {}
            for column in columns:
                value = normalized.get(column)
                if isinstance(value, (dict, list)):
                    row[column] = json.dumps(value, ensure_ascii=False)[:1000]
                elif value is None:
                    row[column] = ""
                else:
                    row[column] = str(value).strip()[:1000]
            rows.append(row)
        return columns, rows

    @staticmethod
    def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                raise ValueError("XLSX worksheet is empty")
            headers = [_header(value) for value in first[:MAX_COLUMNS]]
            if not headers or any(not value for value in headers):
                raise ValueError("XLSX must contain a usable header row")
            if len(set(headers)) != len(headers):
                raise ValueError("XLSX contains duplicate column names after normalisation")
            rows = []
            for values in iterator:
                row = {}
                for index, column in enumerate(headers):
                    value = values[index] if index < len(values) else None
                    row[column] = "" if value is None else str(value).strip()[:1000]
                if any(str(value).strip() for value in row.values()):
                    rows.append(row)
                if len(rows) >= MAX_ROWS:
                    break
            return headers, rows
        finally:
            workbook.close()

    @staticmethod
    def detect_mapping(columns: list[str]) -> dict[str, str]:
        result = {}
        available = set(columns)
        for metric, aliases in ALIASES.items():
            matches = [column for column in columns if column in aliases]
            if matches:
                result[metric] = matches[0]
                continue
            # Conservative partial match only for aliases with 8+ chars, reducing accidental maps.
            for alias in sorted(aliases, key=len, reverse=True):
                if len(alias) >= 8:
                    match = next((column for column in columns if alias in column and column in available), None)
                    if match:
                        result[metric] = match
                        break
        return result

    @classmethod
    def parse(cls, filename: str, content: bytes) -> tuple[str, list[str], list[dict], dict[str, str]]:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in SUPPORTED_SUFFIXES:
            raise ValueError("Structured import must be CSV, JSON or XLSX")
        if suffix == ".csv":
            columns, rows = cls._parse_csv(content)
            source_format = "csv"
        elif suffix == ".json":
            columns, rows = cls._parse_json(content)
            source_format = "json"
        else:
            columns, rows = cls._parse_xlsx(content)
            source_format = "xlsx"
        if not rows:
            raise ValueError("The import contains no data rows")
        return source_format, columns, rows, cls.detect_mapping(columns)

    def stage(self, user_id: str, filename: str, content: bytes, content_type: str = "") -> dict:
        source_format, columns, rows, mapping = self.parse(filename, content)
        safe_name, upload_path = save_progress_upload(user_id, filename, content)
        import_id, now = uuid4().hex, _now()
        with self._connect() as con:
            con.execute(
                """INSERT INTO esp_creator_data_imports
                   (id,user_id,upload_name,upload_path,upload_content_type,source_format,columns_json,rows_json,
                    detected_mapping_json,status,imported_submission_ids_json,created_at,updated_at,imported_at,rejected_at)
                   VALUES (?,?,?,?,?,?,?,?,?,'staged','[]',?,?,NULL,NULL)""",
                (
                    import_id, user_id, safe_name, upload_path, _clean(content_type, 160), source_format,
                    json.dumps(columns), json.dumps(rows, ensure_ascii=False), json.dumps(mapping, sort_keys=True), now, now,
                ),
            )
        return self.get(user_id, import_id)

    @staticmethod
    def _decode(row) -> dict | None:
        if not row:
            return None
        item = dict(row)
        for source, target, fallback in (
            ("columns_json", "columns", []),
            ("rows_json", "rows", []),
            ("detected_mapping_json", "detected_mapping", {}),
            ("imported_submission_ids_json", "imported_submission_ids", []),
        ):
            try:
                item[target] = json.loads(item.pop(source) or json.dumps(fallback))
            except Exception:
                item[target] = fallback
        item.pop("upload_path", None)
        item["row_count"] = len(item["rows"])
        item["preview_rows"] = item["rows"][:5]
        item.pop("rows", None)
        item["private_upload_path_exposed"] = False
        return item

    def get(self, user_id: str, import_id: str) -> dict:
        with self._connect() as con:
            row = con.execute(
                "SELECT * FROM esp_creator_data_imports WHERE id=? AND user_id=?",
                (import_id, user_id),
            ).fetchone()
        if not row:
            raise KeyError(import_id)
        return self._decode(row) or {}

    def _raw(self, user_id: str, import_id: str):
        with self._connect() as con:
            row = con.execute(
                "SELECT * FROM esp_creator_data_imports WHERE id=? AND user_id=?",
                (import_id, user_id),
            ).fetchone()
        if not row:
            raise KeyError(import_id)
        return row

    def list_for_user(self, user_id: str) -> list[dict]:
        with self._connect() as con:
            rows = con.execute(
                "SELECT * FROM esp_creator_data_imports WHERE user_id=? ORDER BY created_at DESC LIMIT 50",
                (user_id,),
            ).fetchall()
        return [self._decode(row) or {} for row in rows]

    def confirm(self, user_id: str, import_id: str, body: ImportConfirm) -> dict:
        raw = self._raw(user_id, import_id)
        if raw["status"] != "staged":
            raise ValueError("This data import has already been resolved")
        columns = json.loads(raw["columns_json"] or "[]")
        rows = json.loads(raw["rows_json"] or "[]")
        mapping = body.mapping or json.loads(raw["detected_mapping_json"] or "{}")
        if not mapping:
            raise ValueError("Map at least one supported progress metric before importing")
        invalid_columns = sorted({column for column in mapping.values() if column not in columns})
        if invalid_columns:
            raise ValueError("One or more mapped source columns do not exist in this import")
        period_column = _header(body.period_column)
        if period_column and period_column not in columns:
            raise ValueError("Selected period column does not exist in this import")
        submission_ids = []
        skipped_rows = 0
        for index, row in enumerate(rows[:MAX_ROWS], 1):
            metrics = {}
            for metric, column in mapping.items():
                value = _number(row.get(column))
                if value is not None:
                    metrics[metric] = value
            if not metrics:
                skipped_rows += 1
                continue
            period = _clean(row.get(period_column), 160) if period_column else ""
            if not period:
                period = _clean(body.default_period_label, 160) or "Imported analytics"
                if len(rows) > 1:
                    period = f"{period} · row {index}"[:160]
            submission = self.progress.add(
                user_id,
                kind=body.kind,
                period_label=period,
                metrics=metrics,
                notes=_clean(body.notes, 2000),
                upload_name=raw["upload_name"],
                upload_path=raw["upload_path"],
                upload_content_type=raw["upload_content_type"],
            )
            submission_ids.append(submission["id"])
        if not submission_ids:
            raise ValueError("No rows contained usable numeric values for the selected metric mapping")
        now = _now()
        with self._connect() as con:
            con.execute(
                """UPDATE esp_creator_data_imports SET status='imported',imported_submission_ids_json=?,updated_at=?,imported_at=?
                   WHERE id=? AND user_id=? AND status='staged'""",
                (json.dumps(submission_ids), now, now, import_id, user_id),
            )
        return {
            "import": self.get(user_id, import_id),
            "imported_rows": len(submission_ids),
            "skipped_rows": skipped_rows,
            "submission_ids": submission_ids,
            "human_confirmation_required": True,
        }

    def reject(self, user_id: str, import_id: str) -> dict:
        raw = self._raw(user_id, import_id)
        if raw["status"] != "staged":
            raise ValueError("This data import has already been resolved")
        now = _now()
        with self._connect() as con:
            con.execute(
                "UPDATE esp_creator_data_imports SET status='rejected',updated_at=?,rejected_at=? WHERE id=? AND user_id=?",
                (now, now, import_id, user_id),
            )
        return self.get(user_id, import_id)


data_imports = CreatorDataImportStore()


@router.get("/command-center/api/progress/imports")
def data_import_list_api(request: Request):
    member = _creator(request)
    return {
        "imports": data_imports.list_for_user(member.user_id),
        "supported_formats": sorted(SUPPORTED_SUFFIXES),
        "standard_metrics": list(STANDARD_METRICS),
        "private_upload_paths_exposed": False,
    }


@router.post("/command-center/api/progress/imports/{import_id}/confirm")
def confirm_data_import_api(import_id: str, body: ImportConfirm, request: Request):
    member = _creator(request)
    try:
        return data_imports.confirm(member.user_id, import_id, body)
    except KeyError as exc:
        raise HTTPException(404, "Creator data import not found") from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post("/command-center/api/progress/imports/{import_id}/reject")
def reject_data_import_api(import_id: str, request: Request):
    member = _creator(request)
    try:
        return {"import": data_imports.reject(member.user_id, import_id)}
    except KeyError as exc:
        raise HTTPException(404, "Creator data import not found") from exc
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc


CSS = """
:root{--line:#ffffff20;--muted:#c8bfd2;--gold:#efc66b;--violet:#a26dff;--green:#78dda7}*{box-sizing:border-box}body{margin:0;background:radial-gradient(circle at 88% 0,#42185d,transparent 31%),#07050d;color:#fff;font-family:Inter,system-ui,sans-serif}.wrap{width:min(1180px,calc(100% - 28px));margin:auto;padding:34px 0 60px}.top,.row{display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap}.eyebrow{color:var(--gold);font-weight:900;letter-spacing:.14em;text-transform:uppercase;font-size:.72rem}h1{font-size:clamp(2.5rem,7vw,5rem);letter-spacing:-.05em;margin:.15em 0}.muted{color:var(--muted);line-height:1.5}.card{border:1px solid var(--line);border-radius:18px;padding:16px;background:#14101deb;margin:12px 0}.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:9px}.btn,button{display:inline-block;border:1px solid var(--line);border-radius:10px;padding:9px 12px;background:#ffffff09;color:#fff;text-decoration:none;font-weight:800;cursor:pointer}.primary{border:0;background:linear-gradient(110deg,var(--gold),var(--violet));color:#160d1d}.pill{border:1px solid var(--line);border-radius:999px;padding:4px 8px;font-size:.72rem}input,select{width:100%;border:1px solid var(--line);border-radius:10px;padding:10px;background:#080610;color:#fff;margin:5px 0}table{width:100%;border-collapse:collapse}th,td{border-bottom:1px solid var(--line);padding:7px;text-align:left;white-space:nowrap}.scroll{overflow:auto}@media(max-width:760px){.grid{grid-template-columns:1fr}}
"""


def _preview_table(item: dict) -> str:
    columns = item.get("columns") or []
    rows = item.get("preview_rows") or []
    if not columns:
        return ""
    head = "".join(f"<th>{escape(column)}</th>" for column in columns[:12])
    body = "".join(
        "<tr>" + "".join(f"<td>{escape(str(row.get(column) or ''))}</td>" for column in columns[:12]) + "</tr>"
        for row in rows
    )
    return f"<div class='scroll'><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"


@router.get("/command-center/progress/import", response_class=HTMLResponse, include_in_schema=False)
def data_import_page(request: Request):
    member = _creator(request)
    imports = data_imports.list_for_user(member.user_id)
    cards = "".join(
        "<article class='card'><div class='row'><div>"
        f"<span class='pill'>{escape(item['status'].upper())}</span><h2>{escape(item['upload_name'])}</h2>"
        f"<p class='muted'>{item['row_count']} rows · {escape(item['source_format'].upper())} · detected: {escape(', '.join(item.get('detected_mapping', {}).keys()) or 'no standard metrics')}</p></div>"
        f"<a class='btn' href='/command-center/progress/import/{escape(item['id'], quote=True)}'>Review</a></div></article>"
        for item in imports
    ) or "<div class='card muted'>No structured analytics imports yet.</div>"
    return HTMLResponse(
        "<!doctype html><html><head><meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<meta name='robots' content='noindex,nofollow'><title>ESP Creator Data Import</title><style>{CSS}</style></head>"
        "<body><main class='wrap'><div class='top'><div><div class='eyebrow'>Elevate Souls Productions · Creator OS</div>"
        "<h1>Performance Data Import</h1><p class='muted'>Preview and map your own analytics export before anything enters Creator Progress.</p></div>"
        "<a class='btn' href='/command-center/progress'>Creator Progress</a></div>"
        f"{cards}<section class='card'><h2>Stage an export</h2><form method='post' action='/command-center/progress/import' enctype='multipart/form-data'>"
        "<input type='file' name='data_file' accept='.csv,.json,.xlsx' required><p class='muted'>CSV, JSON or XLSX up to the existing 10 MB private-progress upload limit. Maximum 500 rows and 50 columns per staged import.</p>"
        "<button class='primary' type='submit'>Upload & preview</button></form></section>"
        "<section class='card'><b>Nothing imports automatically.</b><p class='muted'>Pulsar detects likely columns, but you confirm the mapping first. Original files remain private and raw server paths are never shown in the member API.</p></section>"
        "</main></body></html>",
        headers={"Cache-Control": "no-store"},
    )


@router.post("/command-center/progress/import", include_in_schema=False)
async def stage_data_import_page(request: Request, data_file: UploadFile = File(...)):
    member = _creator(request)
    content = await data_file.read(10 * 1024 * 1024 + 1)
    try:
        item = data_imports.stage(member.user_id, data_file.filename or "analytics.csv", content, data_file.content_type or "")
    except (ValueError, UnicodeError, json.JSONDecodeError) as exc:
        raise HTTPException(400, str(exc)) from exc
    return RedirectResponse(f"/command-center/progress/import/{item['id']}", status_code=303)


@router.get("/command-center/progress/import/{import_id}", response_class=HTMLResponse, include_in_schema=False)
def review_data_import_page(import_id: str, request: Request):
    member = _creator(request)
    try:
        item = data_imports.get(member.user_id, import_id)
    except KeyError as exc:
        raise HTTPException(404, "Creator data import not found") from exc
    mapping = item.get("detected_mapping") or {}
    map_text = ", ".join(f"{metric} ← {column}" for metric, column in mapping.items()) or "No standard metrics auto-detected. Use the API mapping workflow to choose columns before confirmation."
    return HTMLResponse(
        "<!doctype html><html><head><meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<meta name='robots' content='noindex,nofollow'><title>Review Creator Import</title><style>{CSS}</style></head>"
        f"<body><main class='wrap'><div class='top'><div><div class='eyebrow'>Import Preview</div><h1>{escape(item['upload_name'])}</h1>"
        f"<p class='muted'>{item['row_count']} rows · {escape(item['source_format'].upper())} · status {escape(item['status'])}</p></div>"
        "<a class='btn' href='/command-center/progress/import'>All imports</a></div>"
        f"<section class='card'><h2>Detected mapping</h2><p>{escape(map_text)}</p></section>{_preview_table(item)}"
        "<section class='card'><b>Confirmation required</b><p class='muted'>The preview does not change Creator Progress. Confirm through the private import API with live/video type, final metric mapping and optional period column. A rejected or already imported file cannot be imported twice.</p></section>"
        "</main></body></html>",
        headers={"Cache-Control": "no-store"},
    )


__all__ = ["router", "CreatorDataImportStore", "data_imports", "STANDARD_METRICS", "SUPPORTED_SUFFIXES"]
