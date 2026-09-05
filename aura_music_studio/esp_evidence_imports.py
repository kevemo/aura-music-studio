from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, Field
from pypdf import PdfReader

from .audit import AuditLedger
from .esp_niche import require_esp_hub_member
from .esp_product_workflows import (
    EvidenceBatchInput,
    EvidenceMetricInput,
    _authorize_creator_record,
    _roles,
    workflows,
)

router = APIRouter(tags=["Chat 9 Evidence Import Center"])

ImportSource = Literal["screenshot", "csv", "xlsx", "pdf"]
MappingSource = Literal["csv", "xlsx"]

MAX_IMPORT_BYTES = 12 * 1024 * 1024
MAX_TABULAR_ROWS = 500
MAX_TABULAR_COLUMNS = 100
MAX_PREVIEW_ROWS = 50
MAX_PDF_PAGES = 30
MAX_PDF_TEXT = 24000
PREVIEW_TTL_MINUTES = 30

_SHA256 = re.compile(r"^[0-9a-f]{64}$")


def _now_dt() -> datetime:
    return datetime.now(timezone.utc)


def _now() -> str:
    return _now_dt().isoformat()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _loads(value: str | None, default: Any) -> Any:
    try:
        return json.loads(value or "")
    except Exception:
        return default


def _clean(value: str | None, limit: int) -> str:
    return " ".join((value or "").split())[:limit]


def _normalise_metric_name(value: Any) -> str:
    clean = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return clean[:120]


def _json_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()[:2000]


def _coerce_value(value: Any) -> Any:
    value = _json_cell(value)
    if not isinstance(value, str):
        return value
    clean = value.strip()
    if not clean:
        return None
    if re.fullmatch(r"-?\d+", clean):
        try:
            return int(clean)
        except Exception:
            return clean
    if re.fullmatch(r"-?(?:\d+\.\d+|\d+\.\d*|\.\d+)", clean):
        try:
            return float(clean)
        except Exception:
            return clean
    return clean


def _unique_headers(values: list[Any]) -> tuple[list[str], list[str]]:
    headers: list[str] = []
    warnings: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values[:MAX_TABULAR_COLUMNS], start=1):
        base = _clean(str(value or ""), 160) or f"column_{index}"
        key = base.casefold()
        count = seen.get(key, 0) + 1
        seen[key] = count
        if count > 1:
            warnings.append(f"Duplicate header '{base}' was renamed for preview safety.")
            base = f"{base}__{count}"
        headers.append(base)
    return headers, warnings


def _column_lookup(columns: list[str]) -> dict[str, str]:
    return {column.casefold(): column for column in columns}


def _find_column(columns: list[str], aliases: tuple[str, ...]) -> str | None:
    lookup = _column_lookup(columns)
    for alias in aliases:
        found = lookup.get(alias.casefold())
        if found:
            return found
    return None


def _metrics_from_rows(
    columns: list[str],
    rows: list[dict[str, Any]],
    mapping: dict | None = None,
) -> tuple[list[dict[str, Any]], list[str], bool]:
    warnings: list[str] = []
    if mapping:
        metric_column = mapping.get("metric_column") or ""
        value_column = mapping.get("value_column") or ""
        unit_column = mapping.get("unit_column") or ""
        default_unit = mapping.get("default_unit") or ""
        missing = [column for column in (metric_column, value_column) if column not in columns]
        if unit_column and unit_column not in columns:
            missing.append(unit_column)
        if missing:
            return [], [f"Saved mapping references missing column(s): {', '.join(missing)}."], True
    else:
        metric_column = _find_column(columns, ("metric", "metric_name", "metric name", "name")) or ""
        value_column = _find_column(columns, ("value", "metric_value", "metric value")) or ""
        unit_column = _find_column(columns, ("unit", "units")) or ""
        default_unit = ""
        if not metric_column or not value_column:
            return [], [
                "No deterministic metric/value column pair was found. Choose or create a saved mapping before committing structured metrics."
            ], True

    metrics: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        metric_name = _normalise_metric_name(row.get(metric_column))
        if not metric_name:
            continue
        if metric_name in seen:
            warnings.append(
                f"Metric '{metric_name}' appears more than once; later duplicate rows were excluded from candidate metrics."
            )
            continue
        seen.add(metric_name)
        unit = _clean(str(row.get(unit_column) or default_unit), 40) if unit_column else _clean(default_unit, 40)
        metrics.append(
            {
                "name": metric_name,
                "value": _coerce_value(row.get(value_column)),
                "unit": unit,
                "confidence": 1.0,
                "source_row": row_number,
                "review_required_before_commit": True,
            }
        )
    if not metrics:
        warnings.append("The mapped columns did not contain any named metric rows.")
    return metrics[:200], warnings, not bool(metrics)


def _preview_csv(data: bytes, mapping: dict | None = None) -> dict:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 or UTF-8 with BOM") from exc
    if "\x00" in text:
        raise ValueError("CSV contains binary/null data")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        first = next(reader)
    except StopIteration:
        raise ValueError("CSV is empty")
    if len(first) > MAX_TABULAR_COLUMNS:
        raise ValueError(f"CSV has more than {MAX_TABULAR_COLUMNS} columns")
    columns, warnings = _unique_headers(first)
    rows: list[dict[str, Any]] = []
    truncated = False
    for index, values in enumerate(reader):
        if index >= MAX_TABULAR_ROWS:
            truncated = True
            break
        padded = list(values[: len(columns)]) + [""] * max(0, len(columns) - len(values))
        rows.append({columns[i]: _json_cell(padded[i]) for i in range(len(columns))})
    candidates, metric_warnings, mapping_required = _metrics_from_rows(columns, rows, mapping)
    warnings.extend(metric_warnings)
    return {
        "format": "csv",
        "columns": columns,
        "rows": rows[:MAX_PREVIEW_ROWS],
        "row_count_scanned": len(rows),
        "rows_truncated": truncated or len(rows) > MAX_PREVIEW_ROWS,
        "candidate_metrics": candidates,
        "mapping_required": mapping_required,
        "warnings": warnings,
    }


def _preview_xlsx(data: bytes, mapping: dict | None = None) -> dict:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("XLSX could not be opened as a valid workbook") from exc
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        try:
            first = list(next(iterator))
        except StopIteration:
            raise ValueError("XLSX worksheet is empty")
        if len(first) > MAX_TABULAR_COLUMNS:
            raise ValueError(f"XLSX has more than {MAX_TABULAR_COLUMNS} columns")
        columns, warnings = _unique_headers(first)
        rows: list[dict[str, Any]] = []
        truncated = False
        for index, values in enumerate(iterator):
            if index >= MAX_TABULAR_ROWS:
                truncated = True
                break
            values = list(values[: len(columns)])
            padded = values + [None] * max(0, len(columns) - len(values))
            rows.append({columns[i]: _json_cell(padded[i]) for i in range(len(columns))})
        candidates, metric_warnings, mapping_required = _metrics_from_rows(columns, rows, mapping)
        warnings.extend(metric_warnings)
        return {
            "format": "xlsx",
            "worksheet": worksheet.title[:160],
            "columns": columns,
            "rows": rows[:MAX_PREVIEW_ROWS],
            "row_count_scanned": len(rows),
            "rows_truncated": truncated or len(rows) > MAX_PREVIEW_ROWS,
            "candidate_metrics": candidates,
            "mapping_required": mapping_required,
            "warnings": warnings,
        }
    finally:
        workbook.close()


def _preview_pdf(data: bytes) -> dict:
    if not data.startswith(b"%PDF"):
        raise ValueError("File does not have a valid PDF signature")
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:
        raise ValueError("PDF could not be parsed") from exc
    if getattr(reader, "is_encrypted", False):
        try:
            unlocked = reader.decrypt("")
        except Exception:
            unlocked = 0
        if not unlocked:
            raise ValueError("Encrypted PDF cannot be previewed without a password")
    text_parts: list[str] = []
    pages_scanned = min(len(reader.pages), MAX_PDF_PAGES)
    for page in reader.pages[:pages_scanned]:
        try:
            text_parts.append(page.extract_text() or "")
        except Exception:
            text_parts.append("")
        if sum(len(part) for part in text_parts) >= MAX_PDF_TEXT:
            break
    text = "\n".join(text_parts)[:MAX_PDF_TEXT]
    return {
        "format": "pdf",
        "page_count": len(reader.pages),
        "pages_scanned": pages_scanned,
        "text_preview": text,
        "candidate_metrics": [],
        "mapping_required": True,
        "warnings": [
            "PDF text is preview-only. Numeric values are not silently converted into metrics; add or correct structured metrics after human review."
        ],
    }


def _preview_image(data: bytes) -> dict:
    try:
        with Image.open(io.BytesIO(data)) as image:
            image.verify()
        with Image.open(io.BytesIO(data)) as image:
            width, height = image.size
            image_format = (image.format or "").upper()
    except (UnidentifiedImageError, OSError) as exc:
        raise ValueError("Screenshot/image could not be verified") from exc
    return {
        "format": "image",
        "image_format": image_format,
        "width": width,
        "height": height,
        "candidate_metrics": [],
        "mapping_required": True,
        "warnings": [
            "Image evidence is retained for human review. OCR is not auto-committed as metric truth."
        ],
    }


def preview_uploaded_evidence(
    data: bytes,
    *,
    source_type: str,
    filename: str,
    content_type: str,
    mapping: dict | None = None,
) -> dict:
    if not data:
        raise ValueError("Evidence file is empty")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError(f"Evidence file exceeds {MAX_IMPORT_BYTES // (1024 * 1024)} MB")
    suffix = Path(filename or "").suffix.lower()
    content_type = (content_type or "").split(";", 1)[0].strip().lower()

    if source_type == "csv":
        if suffix != ".csv":
            raise ValueError("CSV imports require a .csv file")
        result = _preview_csv(data, mapping)
    elif source_type == "xlsx":
        if suffix != ".xlsx":
            raise ValueError("XLSX imports require a .xlsx file")
        if not data.startswith(b"PK"):
            raise ValueError("XLSX does not have the expected ZIP container signature")
        result = _preview_xlsx(data, mapping)
    elif source_type == "pdf":
        if suffix != ".pdf":
            raise ValueError("PDF imports require a .pdf file")
        result = _preview_pdf(data)
    elif source_type == "screenshot":
        if suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
            raise ValueError("Screenshot imports support PNG, JPEG or WebP")
        if not content_type.startswith("image/"):
            raise ValueError("Screenshot upload must use an image content type")
        result = _preview_image(data)
    else:
        raise ValueError("Unsupported evidence import source")

    result.update(
        {
            "original_filename": Path(filename or "evidence").name[:240],
            "content_type": content_type[:160],
            "file_size": len(data),
            "source_sha256": hashlib.sha256(data).hexdigest(),
            "auto_commit": False,
            "human_review_required": True,
        }
    )
    return result


class MappingTemplateCreate(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    source_type: MappingSource
    metric_column: str = Field(min_length=1, max_length=160)
    value_column: str = Field(min_length=1, max_length=160)
    unit_column: str = Field(default="", max_length=160)
    default_unit: str = Field(default="", max_length=40)


class EvidenceImportCommit(BaseModel):
    preview_id: str = Field(min_length=8, max_length=64)
    selected_metric_names: list[str] = Field(default_factory=list, max_length=200)
    period_start: str | None = Field(default=None, max_length=40)
    period_end: str | None = Field(default=None, max_length=40)
    captured_at: str | None = Field(default=None, max_length=80)
    notes: str = Field(default="", max_length=2000)


class EvidenceImportStore:
    def __init__(self, workflow_store=None):
        self.workflows = workflow_store or workflows
        self.db_path = self.workflows.db_path
        self.audit = AuditLedger(self.workflows.esp.accounts)
        self._init_schema()

    def _connect(self):
        con = sqlite3.connect(self.db_path, timeout=30)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA foreign_keys=ON")
        con.execute("PRAGMA journal_mode=WAL")
        return con

    def _init_schema(self) -> None:
        with self._connect() as con:
            con.executescript(
                """
                CREATE TABLE IF NOT EXISTS esp_evidence_mapping_templates (
                    id TEXT PRIMARY KEY,
                    owner_user_id TEXT NOT NULL,
                    name TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    metric_column TEXT NOT NULL,
                    value_column TEXT NOT NULL,
                    unit_column TEXT NOT NULL DEFAULT '',
                    default_unit TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(owner_user_id,source_type,name),
                    FOREIGN KEY(owner_user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_chat9_evidence_mappings_owner
                    ON esp_evidence_mapping_templates(owner_user_id,source_type,name);

                CREATE TABLE IF NOT EXISTS esp_evidence_import_previews (
                    id TEXT PRIMARY KEY,
                    actor_user_id TEXT NOT NULL,
                    creator_user_id TEXT NOT NULL,
                    source_sha256 TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    provider TEXT NOT NULL DEFAULT '',
                    raw_evidence_ref TEXT NOT NULL,
                    original_filename TEXT NOT NULL,
                    content_type TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    preview_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    expires_at TEXT NOT NULL,
                    FOREIGN KEY(actor_user_id) REFERENCES users(id) ON DELETE CASCADE,
                    FOREIGN KEY(creator_user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_chat9_evidence_previews_actor
                    ON esp_evidence_import_previews(actor_user_id,created_at DESC);

                CREATE TABLE IF NOT EXISTS esp_evidence_import_hashes (
                    creator_user_id TEXT NOT NULL,
                    source_sha256 TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    raw_evidence_ref TEXT NOT NULL,
                    batch_id TEXT,
                    reserved_by TEXT NOT NULL,
                    reserved_at TEXT NOT NULL,
                    imported_at TEXT,
                    PRIMARY KEY(creator_user_id,source_sha256),
                    FOREIGN KEY(creator_user_id) REFERENCES users(id) ON DELETE CASCADE
                );
                """
            )

    def create_mapping(self, owner_user_id: str, payload: MappingTemplateCreate) -> dict:
        now = _now()
        mapping_id = uuid4().hex
        try:
            with self._connect() as con:
                con.execute(
                    """INSERT INTO esp_evidence_mapping_templates
                       (id,owner_user_id,name,source_type,metric_column,value_column,unit_column,default_unit,created_at,updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (
                        mapping_id,
                        owner_user_id,
                        _clean(payload.name, 120),
                        payload.source_type,
                        _clean(payload.metric_column, 160),
                        _clean(payload.value_column, 160),
                        _clean(payload.unit_column, 160),
                        _clean(payload.default_unit, 40),
                        now,
                        now,
                    ),
                )
        except sqlite3.IntegrityError as exc:
            raise FileExistsError("mapping_template_exists") from exc
        self.audit.append(
            actor=owner_user_id,
            action="chat9.evidence_mapping_created",
            details={"mapping_id": mapping_id, "source_type": payload.source_type},
        )
        return self.mapping(mapping_id, owner_user_id)

    def mapping(self, mapping_id: str, owner_user_id: str) -> dict:
        with self._connect() as con:
            row = con.execute(
                "SELECT * FROM esp_evidence_mapping_templates WHERE id=? AND owner_user_id=?",
                (mapping_id, owner_user_id),
            ).fetchone()
        if row is None:
            raise KeyError("Mapping template not found")
        return dict(row)

    def mappings(self, owner_user_id: str, source_type: str | None = None) -> list[dict]:
        with self._connect() as con:
            if source_type:
                rows = con.execute(
                    """SELECT * FROM esp_evidence_mapping_templates
                       WHERE owner_user_id=? AND source_type=? ORDER BY name""",
                    (owner_user_id, source_type),
                ).fetchall()
            else:
                rows = con.execute(
                    "SELECT * FROM esp_evidence_mapping_templates WHERE owner_user_id=? ORDER BY source_type,name",
                    (owner_user_id,),
                ).fetchall()
        return [dict(row) for row in rows]

    def create_preview(
        self,
        *,
        actor_user_id: str,
        creator_user_id: str,
        source_type: str,
        provider: str,
        raw_evidence_ref: str,
        preview: dict,
    ) -> dict:
        if not self.workflows._user_exists(creator_user_id):
            raise KeyError("Creator not found")
        membership = self.workflows.esp.membership(creator_user_id)
        if not membership or membership.get("status") not in {"active", "owner"}:
            raise ValueError("Creator does not have active ESP access")
        if "creator" not in _roles(membership) and "owner" not in _roles(membership):
            raise ValueError("Evidence imports require an ESP Creator target")
        raw_ref = self.workflows._validate_evidence_ref(raw_evidence_ref)
        created = _now_dt()
        preview_id = uuid4().hex
        expires = created + timedelta(minutes=PREVIEW_TTL_MINUTES)
        with self._connect() as con:
            con.execute(
                """INSERT INTO esp_evidence_import_previews
                   (id,actor_user_id,creator_user_id,source_sha256,source_type,provider,raw_evidence_ref,
                    original_filename,content_type,file_size,preview_json,created_at,expires_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    preview_id,
                    actor_user_id,
                    creator_user_id,
                    preview["source_sha256"],
                    source_type,
                    _clean(provider, 80),
                    raw_ref,
                    _clean(preview.get("original_filename"), 240),
                    _clean(preview.get("content_type"), 160),
                    int(preview.get("file_size") or 0),
                    _json(preview),
                    created.isoformat(),
                    expires.isoformat(),
                ),
            )
        return self.preview(preview_id, actor_user_id=actor_user_id)

    def preview(self, preview_id: str, *, actor_user_id: str) -> dict:
        with self._connect() as con:
            row = con.execute(
                "SELECT * FROM esp_evidence_import_previews WHERE id=? AND actor_user_id=?",
                (preview_id, actor_user_id),
            ).fetchone()
        if row is None:
            raise KeyError("Evidence preview not found")
        item = dict(row)
        try:
            expires = datetime.fromisoformat(item["expires_at"].replace("Z", "+00:00"))
        except Exception:
            raise ValueError("Evidence preview expiry is invalid")
        if expires <= _now_dt():
            raise ValueError("Evidence preview has expired; preview the source again")
        item["preview"] = _loads(item.pop("preview_json"), {})
        duplicate = self.committed_source(item["creator_user_id"], item["source_sha256"])
        item["duplicate_source"] = bool(duplicate)
        item["existing_batch_id"] = duplicate.get("batch_id") if duplicate else None
        return item

    def committed_source(self, creator_user_id: str, source_sha256: str) -> dict | None:
        with self._connect() as con:
            row = con.execute(
                """SELECT * FROM esp_evidence_import_hashes
                   WHERE creator_user_id=? AND source_sha256=? AND batch_id IS NOT NULL""",
                (creator_user_id, source_sha256),
            ).fetchone()
        return dict(row) if row else None

    def _reserve_source(
        self,
        *,
        creator_user_id: str,
        source_sha256: str,
        source_type: str,
        raw_evidence_ref: str,
        actor_user_id: str,
    ) -> None:
        if not _SHA256.fullmatch(source_sha256):
            raise ValueError("Invalid source SHA-256")
        now = _now_dt()
        stale_before = now - timedelta(minutes=15)
        with self._connect() as con:
            con.execute("BEGIN IMMEDIATE")
            row = con.execute(
                """SELECT * FROM esp_evidence_import_hashes
                   WHERE creator_user_id=? AND source_sha256=?""",
                (creator_user_id, source_sha256),
            ).fetchone()
            if row:
                if row["batch_id"]:
                    raise FileExistsError("duplicate_import: this source file is already committed for the creator")
                try:
                    reserved_at = datetime.fromisoformat(str(row["reserved_at"]).replace("Z", "+00:00"))
                except Exception:
                    reserved_at = now
                if reserved_at > stale_before:
                    raise FileExistsError("import_in_progress: this source file is already being committed")
                con.execute(
                    """UPDATE esp_evidence_import_hashes
                       SET source_type=?,raw_evidence_ref=?,reserved_by=?,reserved_at=?,imported_at=NULL
                       WHERE creator_user_id=? AND source_sha256=?""",
                    (
                        source_type,
                        raw_evidence_ref,
                        actor_user_id,
                        now.isoformat(),
                        creator_user_id,
                        source_sha256,
                    ),
                )
            else:
                con.execute(
                    """INSERT INTO esp_evidence_import_hashes
                       (creator_user_id,source_sha256,source_type,raw_evidence_ref,batch_id,reserved_by,reserved_at,imported_at)
                       VALUES (?,?,?,?,NULL,?,?,NULL)""",
                    (
                        creator_user_id,
                        source_sha256,
                        source_type,
                        raw_evidence_ref,
                        actor_user_id,
                        now.isoformat(),
                    ),
                )

    def _release_source(self, creator_user_id: str, source_sha256: str, actor_user_id: str) -> None:
        with self._connect() as con:
            con.execute(
                """DELETE FROM esp_evidence_import_hashes
                   WHERE creator_user_id=? AND source_sha256=? AND batch_id IS NULL AND reserved_by=?""",
                (creator_user_id, source_sha256, actor_user_id),
            )

    def _finalize_source(self, creator_user_id: str, source_sha256: str, batch_id: str) -> None:
        with self._connect() as con:
            con.execute(
                """UPDATE esp_evidence_import_hashes SET batch_id=?,imported_at=?
                   WHERE creator_user_id=? AND source_sha256=?""",
                (batch_id, _now(), creator_user_id, source_sha256),
            )

    def commit_preview(
        self,
        *,
        actor_user_id: str,
        body: EvidenceImportCommit,
    ) -> dict:
        item = self.preview(body.preview_id, actor_user_id=actor_user_id)
        preview = item["preview"]
        candidate_by_name = {
            _normalise_metric_name(metric.get("name")): metric
            for metric in preview.get("candidate_metrics") or []
            if _normalise_metric_name(metric.get("name"))
        }
        requested: list[str] = []
        for name in body.selected_metric_names:
            normalised = _normalise_metric_name(name)
            if normalised and normalised not in requested:
                requested.append(normalised)
        missing = [name for name in requested if name not in candidate_by_name]
        if missing:
            raise ValueError(f"Selected metric(s) are not present in the reviewed preview: {', '.join(missing)}")
        metrics = [
            EvidenceMetricInput(
                name=name,
                value=candidate_by_name[name].get("value"),
                unit=_clean(candidate_by_name[name].get("unit"), 40),
                confidence=candidate_by_name[name].get("confidence"),
            )
            for name in requested
        ]

        self._reserve_source(
            creator_user_id=item["creator_user_id"],
            source_sha256=item["source_sha256"],
            source_type=item["source_type"],
            raw_evidence_ref=item["raw_evidence_ref"],
            actor_user_id=actor_user_id,
        )
        try:
            batch = self.workflows.create_evidence(
                item["creator_user_id"],
                EvidenceBatchInput(
                    source_type=item["source_type"],
                    provider=item["provider"],
                    period_start=body.period_start,
                    period_end=body.period_end,
                    captured_at=body.captured_at,
                    raw_evidence_ref=item["raw_evidence_ref"],
                    notes=body.notes,
                    metrics=metrics,
                ),
                uploader_user_id=actor_user_id,
            )
        except Exception:
            self._release_source(item["creator_user_id"], item["source_sha256"], actor_user_id)
            raise
        self._finalize_source(item["creator_user_id"], item["source_sha256"], batch["id"])
        self.audit.append(
            actor=actor_user_id,
            action="chat9.evidence_file_committed",
            subject_user_id=item["creator_user_id"],
            details={
                "batch_id": batch["id"],
                "source_type": item["source_type"],
                "selected_metric_count": len(metrics),
                "source_sha256": item["source_sha256"],
            },
        )
        return batch


imports = EvidenceImportStore()


@router.get("/command-center/api/workflows/evidence/mappings")
def list_mapping_templates(request: Request, source_type: str | None = None):
    member, _membership = require_esp_hub_member(request)
    if source_type and source_type not in {"csv", "xlsx"}:
        raise HTTPException(400, "source_type must be csv or xlsx")
    return {"mappings": imports.mappings(member.user_id, source_type)}


@router.post("/command-center/api/workflows/evidence/mappings")
def create_mapping_template(body: MappingTemplateCreate, request: Request):
    member, _membership = require_esp_hub_member(request)
    try:
        return {"mapping": imports.create_mapping(member.user_id, body)}
    except FileExistsError as exc:
        raise HTTPException(409, {"code": "mapping_template_exists"}) from exc


@router.post("/command-center/api/workflows/evidence/import-preview")
async def preview_evidence_import(
    request: Request,
    file: UploadFile = File(...),
    creator_user_id: str = Form(""),
    source_type: ImportSource = Form(...),
    provider: str = Form(""),
    raw_evidence_ref: str = Form(...),
    mapping_template_id: str = Form(""),
):
    member, membership = require_esp_hub_member(request)
    creator_id = creator_user_id.strip() or member.user_id
    _authorize_creator_record(member.user_id, membership, creator_id)
    mapping = None
    if mapping_template_id.strip():
        try:
            mapping = imports.mapping(mapping_template_id.strip(), member.user_id)
        except KeyError as exc:
            raise HTTPException(404, str(exc)) from exc
        if mapping["source_type"] != source_type:
            raise HTTPException(400, "Saved mapping source type does not match uploaded file type")
    data = await file.read(MAX_IMPORT_BYTES + 1)
    try:
        parsed = preview_uploaded_evidence(
            data,
            source_type=source_type,
            filename=file.filename or "evidence",
            content_type=file.content_type or "",
            mapping=mapping,
        )
        preview = imports.create_preview(
            actor_user_id=member.user_id,
            creator_user_id=creator_id,
            source_type=source_type,
            provider=provider,
            raw_evidence_ref=raw_evidence_ref,
            preview=parsed,
        )
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        status = 413 if "exceeds" in str(exc) else 400
        raise HTTPException(status, str(exc)) from exc
    return {
        "preview": preview,
        "direct_backstage_connection": False,
        "auto_committed": False,
        "human_review_required": True,
    }


@router.post("/command-center/api/workflows/evidence/import-commit")
def commit_evidence_import(body: EvidenceImportCommit, request: Request):
    member, membership = require_esp_hub_member(request)
    try:
        preview = imports.preview(body.preview_id, actor_user_id=member.user_id)
        _authorize_creator_record(member.user_id, membership, preview["creator_user_id"])
        batch = imports.commit_preview(actor_user_id=member.user_id, body=body)
    except KeyError as exc:
        raise HTTPException(404, str(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(403, str(exc)) from exc
    except FileExistsError as exc:
        raise HTTPException(409, {"code": "duplicate_import", "message": str(exc)}) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {
        "batch": batch,
        "source_deduplicated": True,
        "direct_backstage_connection": False,
    }


__all__ = [
    "EvidenceImportStore",
    "EvidenceImportCommit",
    "MappingTemplateCreate",
    "preview_uploaded_evidence",
    "imports",
    "router",
]
